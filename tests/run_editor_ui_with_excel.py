from __future__ import annotations

import os
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Browser, Page, sync_playwright


BASE_URL = os.getenv("BASE_URL", "http://127.0.0.1:3999")
SCENARIO_FILE = Path.cwd() / "test_scenrios.txt"
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
RESULTS_FILE = Path.cwd() / f"test_results_{RUN_TIMESTAMP}.xlsx"


def load_scenarios(file_path: Path) -> list[str]:
    if not file_path.exists():
        raise FileNotFoundError(f"Scenario file not found: {file_path}")

    scenarios = [line.strip() for line in file_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    if not scenarios:
        raise ValueError(f"No scenarios found in: {file_path}")
    return scenarios


def write_result_to_excel(result: dict[str, str | int]) -> None:
    if RESULTS_FILE.exists():
        workbook = load_workbook(RESULTS_FILE)
        worksheet = workbook["Results"] if "Results" in workbook.sheetnames else workbook.create_sheet("Results")
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Results"
        worksheet.append(
            [
                "Scenario Name",
                "Date & Time",
                "Status",
                "Execution Time (ms)",
                "Error Message",
            ]
        )

    worksheet.append(
        [
            result["scenario_name"],
            result["executed_at"],
            result["status"],
            result["duration"],
            result["error_message"],
        ]
    )
    workbook.save(RESULTS_FILE)
    print(f"Results saved to Excel: {RESULTS_FILE}")


def validate_visible(page: Page, selector: str, label: str) -> None:
    page.locator(selector).wait_for(state="visible", timeout=15_000)
    print(f"Validated: {label}")


def open_fluent(page: Page) -> None:
    page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30_000)
    page.wait_for_load_state("networkidle")
    print(f"Page loaded: {BASE_URL}")

    validate_visible(page, "#title", "application title")
    validate_visible(page, ".note-editable", "editor area")
    validate_visible(page, "#update_button", "update highlighting button")
    validate_visible(page, "#active_learning", "active learning button")
    validate_visible(page, "#preferences", "preferences button")

    title_text = page.locator("#title").inner_text().strip()
    if title_text != "Fluent":
        raise AssertionError(f'Expected title to be "Fluent" but found "{title_text}"')


def set_editor_text(page: Page, text: str) -> None:
    editor = page.locator(".note-editable")
    editor.click()
    page.keyboard.press("Control+A")
    page.keyboard.type(text)


def close_modal_if_open(page: Page, selector: str) -> None:
    modal = page.locator(selector)
    if modal.count() and modal.is_visible():
        page.locator(f"{selector} .close").first.click()
        page.wait_for_timeout(500)


def run_editor_input_flow(page: Page) -> None:
    set_editor_text(page, "Graph ideas improve confidence.")
    print("Entered short practice sentence into the editor")


def run_default_content_validation_flow(page: Page) -> None:
    editor_text = page.locator(".note-editable").inner_text().strip()
    if not editor_text:
        raise AssertionError("Expected the editor to contain visible content after page load.")
    print("Validated that the editor contains visible content after initial load")


def run_default_smoke_flow(page: Page) -> None:
    set_editor_text(page, "Graph ideas improve confidence.")
    print("Entered practice sentence into the editor")
    page.locator("#update_button").click()
    page.wait_for_timeout(1500)
    print("Triggered update highlighting")


def run_revised_sentence_flow(page: Page) -> None:
    set_editor_text(page, "A crisis can feel grave.")
    print("Entered revised sentence into the editor")
    page.locator("#update_button").click()
    page.wait_for_timeout(1500)
    print("Triggered highlighting for the revised sentence")


def run_paragraph_review_flow(page: Page) -> None:
    set_editor_text(page, "Graph ideas help the group stay calm. A grave tone can make the opening feel harder.")
    print("Entered longer practice paragraph into the editor")
    page.locator("#update_button").click()
    page.wait_for_timeout(1500)
    print("Triggered highlighting for the longer paragraph")


def run_open_preferences_flow(page: Page) -> None:
    page.locator("#preferences").click()
    validate_visible(page, "#preferences_modal", "preferences modal")
    validate_visible(page, "#easy_words", "easy words field")
    validate_visible(page, "#diff_words", "difficult words field")
    print("Opened preferences modal successfully")


def run_preferences_flow(page: Page) -> None:
    run_default_smoke_flow(page)
    run_open_preferences_flow(page)
    page.locator("#threshold").evaluate(
        """
        (element) => {
            element.value = '85';
            element.dispatchEvent(new Event('input', { bubbles: true }));
        }
        """
    )
    page.locator("#update").click()
    page.wait_for_timeout(1000)
    print("Updated preferences and closed modal")


def run_lower_threshold_flow(page: Page) -> None:
    run_default_smoke_flow(page)
    run_open_preferences_flow(page)
    page.locator("#threshold").evaluate(
        """
        (element) => {
            element.value = '60';
            element.dispatchEvent(new Event('input', { bubbles: true }));
        }
        """
    )
    page.locator("#update").click()
    page.wait_for_timeout(1000)
    print("Updated preferences with a lower threshold and closed modal")


def run_open_active_learning_flow(page: Page) -> None:
    page.locator("#active_learning").click()
    validate_visible(page, "#al_modal", "active learning modal")
    validate_visible(page, "#al_word", "active learning word")
    print("Opened active learning modal successfully")


def run_active_learning_flow(page: Page) -> None:
    run_default_smoke_flow(page)
    run_open_active_learning_flow(page)
    close_modal_if_open(page, "#al_modal")
    print("Reviewed active learning prompt")


def run_repeated_update_flow(page: Page) -> None:
    set_editor_text(page, "Graph ideas improve confidence.")
    page.locator("#update_button").click()
    page.wait_for_timeout(1200)
    set_editor_text(page, "Graph ideas build calm confidence.")
    page.locator("#update_button").click()
    page.wait_for_timeout(1200)
    print("Completed repeated update flow")


def run_preferences_return_flow(page: Page) -> None:
    run_open_preferences_flow(page)
    close_modal_if_open(page, "#preferences_modal")
    page.locator(".note-editable").click()
    print("Returned to the editor after opening preferences")


def run_active_learning_return_flow(page: Page) -> None:
    run_open_active_learning_flow(page)
    close_modal_if_open(page, "#al_modal")
    page.locator(".note-editable").click()
    print("Returned to the editor after opening active learning")


def run_scenario_steps(page: Page, scenario_name: str) -> None:
    scenario_key = scenario_name.lower()

    if "type" in scenario_key or "short practice sentence" in scenario_key:
        run_editor_input_flow(page)
        return

    if "load the built-in sample" in scenario_key or "first visit" in scenario_key:
        run_default_content_validation_flow(page)
        return

    if "longer practice paragraph" in scenario_key or "paragraph" in scenario_key:
        run_paragraph_review_flow(page)
        return

    if "revising the current sentence" in scenario_key or "revised sentence" in scenario_key:
        run_revised_sentence_flow(page)
        return

    if "open the preferences modal" in scenario_key:
        run_open_preferences_flow(page)
        return

    if "lower the confidence threshold" in scenario_key or "lower" in scenario_key:
        run_lower_threshold_flow(page)
        return

    if "raise the confidence threshold" in scenario_key or "threshold" in scenario_key:
        run_preferences_flow(page)
        return

    if "open active learning and return" in scenario_key:
        run_active_learning_return_flow(page)
        return

    if "open the active learning modal" in scenario_key:
        run_open_active_learning_flow(page)
        return

    if "active learning" in scenario_key or "learning" in scenario_key:
        run_active_learning_flow(page)
        return

    if "preferences and return" in scenario_key:
        run_preferences_return_flow(page)
        return

    if "multiple times" in scenario_key or "repeated" in scenario_key:
        run_repeated_update_flow(page)
        return

    if "presentation opening line" in scenario_key:
        run_default_smoke_flow(page)
        return

    if "highlight" in scenario_key or "update" in scenario_key:
        run_default_smoke_flow(page)
        return

    if "open" in scenario_key or "startup" in scenario_key or "launch" in scenario_key:
        print("Startup validation completed")
        return

    run_default_smoke_flow(page)


def execute_scenario(browser: Browser, scenario_name: str) -> None:
    print(f"Scenario started: {scenario_name}")
    started_at = time.perf_counter()
    executed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    status = "PASS"
    error_message = ""
    context = None

    try:
        context = browser.new_context(viewport={"width": 1600, "height": 900}, ignore_https_errors=True)
        page = context.new_page()
        open_fluent(page)
        run_scenario_steps(page, scenario_name)
        print(f"Scenario completed: {scenario_name}")
    except Exception as exc:
        status = "FAIL"
        error_message = str(exc)
        print(f"Scenario failed: {scenario_name} | {error_message}")
    finally:
        duration = int((time.perf_counter() - started_at) * 1000)
        write_result_to_excel(
            {
                "scenario_name": scenario_name,
                "executed_at": executed_at,
                "status": status,
                "duration": duration,
                "error_message": error_message,
            }
        )

        if context is not None:
            context.close()


def run_all_scenarios() -> None:
    scenarios = load_scenarios(SCENARIO_FILE)
    browser = None

    print(f"Loaded {len(scenarios)} scenarios from {SCENARIO_FILE}")

    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=False, args=["--start-maximized"])
            print("Browser launched")

            for scenario_name in scenarios:
                execute_scenario(browser, scenario_name)
    finally:
        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass


if __name__ == "__main__":
    run_all_scenarios()
